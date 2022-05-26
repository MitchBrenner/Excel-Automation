import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)  # loads in excel project
    sheet = wb['Sheet1']  # gets 'sheet1' from excel
    # cell = sheet['a1']  # accessing specific cells as coords or cell = sheet.cell(1,1)
    title_cell = sheet.cell(1, 4)
    title_cell.value = "Corrected Price"
    for row in range(2, sheet.max_row + 1):  # generate numbers from 1 to how many rows they have, #start from two to
        # ignore header
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9  # correct the cell value to 0.9 of the price
        corrected_price_cell = sheet.cell(row, 4)  # find new cell that the corrected price will be stored to
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=3,
                       max_col=4)  # this creates a reference of values from row 2 to 4 and col 4 to 4

    chart = BarChart()  # create a new BarChart
    chart.add_data(values)  # add the data from values to the chart
    sheet.add_chart(chart, 'f2')  # add the chart to the sheet at specified place, top left in e2

    wb.save(filename)  # save workbook to new file incase there is a bug


# process_workbook('transactions.xlsx')
process_workbook('transactions2.xlsx')
